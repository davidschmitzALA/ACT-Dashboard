"""
ALA Strategic Plan Work Plan Generator

Config keys
-----------
client_name     : str
project_title   : str  (default "Strategic Plan")
team            : list of {name, role, rate}
num_months      : int
start_month     : "YYYY-MM" or null
admin_fee_pct   : float  (default 3.0, applied to labor only)
travel          : {num_visits, consultants: [{name, airfare_per_visit,
                   housing_per_visit, meals_per_visit}]}
sections        : list of section dicts
"""

import sys
import json
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _thin(top=False, bottom=False, left=False, right=False):
    s = lambda on: Side(style="thin") if on else Side(style=None)
    return Border(top=s(top), bottom=s(bottom), left=s(left), right=s(right))


def _style(cell, bold=False, italic=False, size=10, color="000000",
           bg=None, align="left", valign="center", wrap=False, num_fmt=None):
    cell.font = Font(name="Arial", bold=bold, italic=italic, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt


def build_workplan(config, output_path):
    if isinstance(config, str):
        with open(config) as f:
            cfg = json.load(f)
    else:
        cfg = config

    client_name   = cfg["client_name"]
    project_title = cfg.get("project_title", "Strategic Plan")
    team          = cfg["team"]
    num_months    = cfg["num_months"]
    start_month   = cfg.get("start_month")
    admin_fee_pct = float(cfg.get("admin_fee_pct", 3.0))
    travel_cfg    = cfg.get("travel")
    sections      = cfg["sections"]

    if start_month:
        start_dt = datetime.strptime(start_month, "%Y-%m")
    else:
        start_dt = datetime.today().replace(day=1)

    month_labels = [
        (start_dt + relativedelta(months=i)).strftime("%b '%y")
        for i in range(num_months)
    ]

    # ── Palette ──────────────────────────────────────────────────────────────
    DARK_BLUE  = "1F3864"
    MED_BLUE   = "2E75B6"
    LIGHT_BLUE = "D6E4F0"
    SECTION_BG = "E9F0F8"
    TOTAL_BG   = "D9E1F2"
    CONSULT_BG = "1A5276"
    PROJECT_BG = "0D2137"
    INPUT_BG   = "FFF8DC"
    WHITE      = "FFFFFF"
    LT_GRAY    = "F2F2F2"

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Plan"

    last_col = get_column_letter(4 + num_months)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    for i in range(num_months):
        ws.column_dimensions[get_column_letter(5 + i)].width = 9

    # ── Title header ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"Amplify Leadership Advisors  |  {client_name} — {project_title}"
    _style(c, bold=True, size=13, color=WHITE, bg=DARK_BLUE, align="center")
    ws.row_dimensions[2].height = 6

    # ── Team / rate block ─────────────────────────────────────────────────────
    TEAM_START = 3
    team_row_map = {}

    ws.row_dimensions[TEAM_START].height = 16
    for label, col in [("Team Member", 1), ("Role", 2), ("Rate ($/hr)", 3)]:
        _style(ws.cell(row=TEAM_START, column=col, value=label),
               bold=True, color=WHITE, bg=MED_BLUE, align="center")

    for i, member in enumerate(team):
        r = TEAM_START + 1 + i
        ws.row_dimensions[r].height = 15
        team_row_map[member["name"]] = r
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        _style(ws.cell(row=r, column=1, value=member["name"]), bg=bg)
        _style(ws.cell(row=r, column=2, value=member.get("role", "")), bg=bg)
        _style(ws.cell(row=r, column=3, value=member["rate"]),
               bg=bg, align="center", num_fmt="$#,##0")

    TEAM_END = TEAM_START + len(team)

    # ── WBS column header row ─────────────────────────────────────────────────
    WBS_HDR = TEAM_END + 3
    ws.row_dimensions[WBS_HDR].height = 18
    for col_idx, h in enumerate(
        ["Task / Activity", "Team Member", "Total Cost ($)", "Total Hours"]
        + month_labels, start=1
    ):
        c = ws.cell(row=WBS_HDR, column=col_idx, value=h)
        _style(c, bold=True, color=WHITE, bg=MED_BLUE, align="center", wrap=True)
        c.border = _thin(bottom=True)

    # ── State ─────────────────────────────────────────────────────────────────
    current_row = WBS_HDR + 1
    person_cost_cells  = {m["name"]: [] for m in team}
    person_hours_cells = {m["name"]: [] for m in team}

    # ── Writers ───────────────────────────────────────────────────────────────

    def write_section_header(label):
        nonlocal current_row
        ws.row_dimensions[current_row].height = 16
        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
        c = ws.cell(row=current_row, column=1, value=label)
        _style(c, bold=True, color=DARK_BLUE, bg=SECTION_BG, wrap=True)
        c.border = _thin(top=True, bottom=True)
        current_row += 1

    def write_task_rows(task_label, member_hours_by_name, recurring=False):
        nonlocal current_row
        first = True
        for m_name, monthly_hours in member_hours_by_name.items():
            if m_name not in team_row_map:
                continue
            rate_ref = f"$C${team_row_map[m_name]}"
            r = current_row
            ws.row_dimensions[r].height = 15
            bg = LT_GRAY if r % 2 == 0 else WHITE

            _style(ws.cell(row=r, column=1, value=task_label if first else None),
                   wrap=True, bg=bg)
            _style(ws.cell(row=r, column=2, value=m_name), bg=bg)

            hrs_range = (f"{get_column_letter(5)}{r}:"
                         f"{get_column_letter(4 + num_months)}{r}")
            _style(ws.cell(row=r, column=4, value=f"=SUM({hrs_range})"),
                   bg=bg, align="center", num_fmt="0.0")
            _style(ws.cell(row=r, column=3, value=f"={rate_ref}*D{r}"),
                   bg=bg, align="center", num_fmt="$#,##0")

            # Find last non-zero month so recurring formula stops there
            last_nz = max(
                (i for i, h in enumerate(monthly_hours) if h != 0), default=-1
            )
            for col_i, hrs in enumerate(monthly_hours):
                # Recurring: col E = literal value; F onwards = =$E{row}
                if recurring and col_i > 0 and col_i <= last_nz:
                    value = f"=$E{r}"
                else:
                    value = hrs if hrs else None
                _style(ws.cell(row=r, column=5 + col_i, value=value),
                       align="center", bg=bg, num_fmt="0.0")

            person_cost_cells[m_name].append(f"C{r}")
            person_hours_cells[m_name].append(f"D{r}")
            first = False
            current_row += 1

    def write_interview_task(task_label, default_num=15, default_hrs=1.5):
        """
        Writes two yellow input rows (# interviews, hours/interview) then
        one cost row per team member whose Total Hours = input_n × input_h.
        Month columns are left blank for the user to distribute manually.
        """
        nonlocal current_row

        def _input_row(label, value, fmt="0"):
            nonlocal current_row
            row = current_row
            ws.row_dimensions[row].height = 15
            _style(ws.cell(row=row, column=1, value=f"  {label}"),
                   italic=True, color="555555", bg=INPUT_BG)
            c_b = ws.cell(row=row, column=2, value=value)
            _style(c_b, bold=True, bg=INPUT_BG, align="center", num_fmt=fmt)
            for col in range(3, 5 + num_months):
                _style(ws.cell(row=row, column=col), bg=INPUT_BG)
            current_row += 1
            return row

        n_row = _input_row("# of Interviews  ← enter here", default_num, "0")
        h_row = _input_row("Hours per Interview  ← enter here", default_hrs, "0.0")

        # Small spacer before consultant rows
        ws.row_dimensions[current_row].height = 4
        current_row += 1

        first = True
        for member in team:
            m_name = member["name"]
            if m_name not in team_row_map:
                continue
            rate_ref = f"$C${team_row_map[m_name]}"
            r = current_row
            ws.row_dimensions[r].height = 15
            bg = LT_GRAY if r % 2 == 0 else WHITE

            _style(ws.cell(row=r, column=1, value=task_label if first else None),
                   wrap=True, bg=bg)
            _style(ws.cell(row=r, column=2, value=m_name), bg=bg)

            # D = n × h  (not a SUM — user fills month columns to match)
            _style(ws.cell(row=r, column=4,
                           value=f"=$B${n_row}*$B${h_row}"),
                   bg=bg, align="center", num_fmt="0.0")
            _style(ws.cell(row=r, column=3, value=f"={rate_ref}*D{r}"),
                   bg=bg, align="center", num_fmt="$#,##0")

            # Month columns blank
            for col_i in range(num_months):
                _style(ws.cell(row=r, column=5 + col_i),
                       align="center", bg=bg, num_fmt="0.0")

            person_cost_cells[m_name].append(f"C{r}")
            person_hours_cells[m_name].append(f"D{r}")
            first = False
            current_row += 1

    def write_subtotal(label, row_start, row_end):
        nonlocal current_row
        ws.row_dimensions[current_row].height = 15
        _style(ws.cell(row=current_row, column=1, value=label),
               bold=True, bg=TOTAL_BG)
        _style(ws.cell(row=current_row, column=2), bold=True, bg=TOTAL_BG)
        _style(ws.cell(row=current_row, column=3,
                       value=f"=SUM(C{row_start}:C{row_end})"),
               bold=True, bg=TOTAL_BG, align="center", num_fmt="$#,##0")
        _style(ws.cell(row=current_row, column=4,
                       value=f"=SUM(D{row_start}:D{row_end})"),
               bold=True, bg=TOTAL_BG, align="center", num_fmt="0.0")
        for col_i in range(num_months):
            cl = get_column_letter(5 + col_i)
            _style(ws.cell(row=current_row, column=5 + col_i,
                           value=f"=SUM({cl}{row_start}:{cl}{row_end})"),
                   bold=True, bg=TOTAL_BG, align="center", num_fmt="0.0")
        current_row += 1

    # ── Write sections ────────────────────────────────────────────────────────
    for section in sections:
        write_section_header(section["label"])
        sec_start = current_row
        tasks = section["tasks"]
        for i, task in enumerate(tasks):
            if task.get("type") == "interview":
                write_interview_task(
                    task["label"],
                    task.get("default_num_interviews", 15),
                    task.get("default_hours_per", 1.5),
                )
            else:
                write_task_rows(
                    task["label"],
                    task["members"],
                    recurring=task.get("recurring", False),
                )
            if i < len(tasks) - 1:
                ws.row_dimensions[current_row].height = 5
                current_row += 1
        write_subtotal(f"Subtotal — {section['label']}", sec_start, current_row - 1)
        current_row += 1

    # ── Additional Work sections (3 blank, formula-ready) ─────────────────────
    for add_i in range(1, 4):
        write_section_header(f"Additional Work {add_i}")
        sec_start = current_row
        members = {m["name"]: [0.0] * num_months for m in team}
        write_task_rows(f"Additional Work {add_i}", members)
        write_subtotal(f"Subtotal — Additional Work {add_i}",
                       sec_start, current_row - 1)
        current_row += 1

    # ── Travel section ────────────────────────────────────────────────────────
    travel_cost_cells = []
    if travel_cfg:
        trips             = travel_cfg.get("trips", [])
        housing_per_night = float(travel_cfg.get("housing_per_night", 200))
        meals_per_day     = float(travel_cfg.get("meals_per_day", 75))

        if trips:
            write_section_header("Travel")

            for ti, trip in enumerate(trips):
                trip_label    = trip.get("label", f"Trip {ti + 1}")
                nights        = int(trip.get("nights", 1))
                meal_days     = int(trip.get("meal_days", 1))
                travelers     = trip.get("travelers", [])

                # Trip sub-header row
                nights_str    = f"{nights} night{'s' if nights != 1 else ''}"
                meal_days_str = f"{meal_days} meal day{'s' if meal_days != 1 else ''}"
                ws.row_dimensions[current_row].height = 14
                ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
                c = ws.cell(row=current_row, column=1,
                            value=f"  {trip_label}  ({nights_str}, {meal_days_str})")
                _style(c, bold=True, italic=True, size=9,
                       color=DARK_BLUE, bg=LIGHT_BLUE)
                current_row += 1

                trip_start = current_row
                for ci, traveler in enumerate(travelers):
                    name          = traveler.get("name", "")
                    airfare       = float(traveler.get("airfare", 0))
                    housing_total = nights * housing_per_night
                    meals_total   = meal_days * meals_per_day

                    for label, amount in [
                        (f"    {name} — Airfare", airfare),
                        (f"    {name} — Housing  "
                         f"({nights} nights × ${housing_per_night:,.0f}/night)",
                         housing_total),
                        (f"    {name} — Meals & Incidentals  "
                         f"({meal_days} days × ${meals_per_day:,.0f}/day)",
                         meals_total),
                    ]:
                        r = current_row
                        ws.row_dimensions[r].height = 15
                        bg = LT_GRAY if r % 2 == 0 else WHITE
                        _style(ws.cell(row=r, column=1, value=label), bg=bg)
                        _style(ws.cell(row=r, column=2), bg=bg)
                        _style(ws.cell(row=r, column=3, value=amount),
                               bg=bg, align="center", num_fmt="$#,##0")
                        _style(ws.cell(row=r, column=4), bg=bg)
                        for col_i in range(num_months):
                            _style(ws.cell(row=r, column=5 + col_i), bg=bg)
                        travel_cost_cells.append(f"C{r}")
                        current_row += 1

                    if ci < len(travelers) - 1:
                        ws.row_dimensions[current_row].height = 4
                        current_row += 1

                # Per-trip subtotal
                ws.row_dimensions[current_row].height = 15
                _style(ws.cell(row=current_row, column=1,
                               value=f"Subtotal — {trip_label}"),
                       bold=True, bg=TOTAL_BG)
                _style(ws.cell(row=current_row, column=2), bold=True, bg=TOTAL_BG)
                _style(ws.cell(row=current_row, column=3,
                               value=f"=SUM(C{trip_start}:C{current_row - 1})"),
                       bold=True, bg=TOTAL_BG, align="center", num_fmt="$#,##0")
                for col in range(4, 5 + num_months):
                    _style(ws.cell(row=current_row, column=col),
                           bold=True, bg=TOTAL_BG)
                current_row += 1

                if ti < len(trips) - 1:
                    ws.row_dimensions[current_row].height = 5
                    current_row += 1

            current_row += 1  # gap after last trip block

    # ── Summary ───────────────────────────────────────────────────────────────
    all_cost_cells  = [c for m in team for c in person_cost_cells[m["name"]]]
    all_hours_cells = [c for m in team for c in person_hours_cells[m["name"]]]

    def _summary_row(label, formula, bg, size=11, hours_formula=None):
        nonlocal current_row
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(f"A{current_row}:B{current_row}")
        _style(ws.cell(row=current_row, column=1, value=label),
               bold=True, size=size, color=WHITE, bg=bg)
        _style(ws.cell(row=current_row, column=3, value=formula),
               bold=True, size=size, color=WHITE, bg=bg,
               align="center", num_fmt="$#,##0")
        if hours_formula:
            _style(ws.cell(row=current_row, column=4, value=hours_formula),
                   bold=True, size=size, color=WHITE, bg=bg,
                   align="center", num_fmt="0.0")
            for col_i in range(num_months):
                cl = get_column_letter(5 + col_i)
                _style(ws.cell(row=current_row, column=5 + col_i,
                               value=f"=SUM({cl}{WBS_HDR+1}:{cl}{current_row-1})"),
                       bold=True, color=WHITE, bg=bg,
                       align="center", num_fmt="0.0")
        ref = f"C{current_row}"
        current_row += 1
        return ref

    labor_cost_formula  = ("=SUM(" + ",".join(all_cost_cells) + ")"
                           if all_cost_cells else "=0")
    labor_hours_formula = ("=SUM(" + ",".join(all_hours_cells) + ")"
                           if all_hours_cells else "=0")

    labor_cell = _summary_row("Total Labor", labor_cost_formula, DARK_BLUE,
                               hours_formula=labor_hours_formula)

    # Admin fee — percentage lives in a yellow col-D cell so it's easy to change
    admin_row_num = current_row
    ws.row_dimensions[admin_row_num].height = 20
    ws.merge_cells(f"A{admin_row_num}:B{admin_row_num}")
    _style(ws.cell(row=admin_row_num, column=1, value="Administrative Fee"),
           bold=True, size=10, color=WHITE, bg=DARK_BLUE)
    _style(ws.cell(row=admin_row_num, column=3,
                   value=f"={labor_cell}*D{admin_row_num}"),
           bold=True, size=10, color=WHITE, bg=DARK_BLUE,
           align="center", num_fmt="$#,##0")
    _style(ws.cell(row=admin_row_num, column=4,
                   value=admin_fee_pct / 100),
           bold=True, size=10, bg="FFFF00",
           align="center", num_fmt="0.0%")
    admin_cell = f"C{admin_row_num}"
    current_row += 1
    consult_cell  = _summary_row(
        "Consulting Fee  (labor + admin)",
        f"={labor_cell}+{admin_cell}",
        CONSULT_BG,
    )

    travel_total_cell = None
    if travel_cost_cells:
        travel_total_cell = _summary_row(
            "Total Travel",
            "=SUM(" + ",".join(travel_cost_cells) + ")",
            DARK_BLUE, size=10,
        )

    total_parts = [consult_cell] + ([travel_total_cell] if travel_total_cell else [])
    _summary_row(
        "PROJECT TOTAL",
        "=SUM(" + ",".join(total_parts) + ")",
        PROJECT_BG, size=12,
    )
    current_row += 1

    # ── Per-person summary ────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 16
    for col_idx, label in enumerate(
        ["Team Member", "Role", "Total Cost", "Total Hours"], 1
    ):
        _style(ws.cell(row=current_row, column=col_idx, value=label),
               bold=True, color=WHITE, bg=MED_BLUE, align="center")
    current_row += 1

    for i, member in enumerate(team):
        m_name = member["name"]
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws.row_dimensions[current_row].height = 15
        _style(ws.cell(row=current_row, column=1, value=m_name), bg=bg)
        _style(ws.cell(row=current_row, column=2,
                       value=member.get("role", "")), bg=bg)
        cc = person_cost_cells[m_name]
        _style(ws.cell(row=current_row, column=3,
                       value="=SUM(" + ",".join(cc) + ")" if cc else 0),
               bg=bg, align="center", num_fmt="$#,##0")
        hc = person_hours_cells[m_name]
        _style(ws.cell(row=current_row, column=4,
                       value="=SUM(" + ",".join(hc) + ")" if hc else 0),
               bg=bg, align="center", num_fmt="0.0")
        current_row += 1

    # ── Payment Schedule ──────────────────────────────────────────────────────
    current_row += 2

    ws.merge_cells(f"A{current_row}:D{current_row}")
    _style(ws.cell(row=current_row, column=1, value="PAYMENT SCHEDULE"),
           bold=True, size=12, color=WHITE, bg=DARK_BLUE, align="center")
    current_row += 1

    # Consulting fee reference row
    ws.merge_cells(f"A{current_row}:B{current_row}")
    _style(ws.cell(row=current_row, column=1,
                   value="Consulting Fee (labor + admin, excl. travel)"),
           bold=True, bg=LIGHT_BLUE)
    _style(ws.cell(row=current_row, column=3, value=f"={consult_cell}"),
           bold=True, bg=LIGHT_BLUE, align="center", num_fmt="$#,##0")
    _style(ws.cell(row=current_row, column=4), bg=LIGHT_BLUE)
    pay_base = f"C{current_row}"
    current_row += 1

    # Column headers
    for col_idx, label in enumerate(["Invoice", "Due Date", "Amount"], 1):
        _style(ws.cell(row=current_row, column=col_idx, value=label),
               bold=True, color=WHITE, bg=MED_BLUE, align="center")
    _style(ws.cell(row=current_row, column=4), bg=MED_BLUE)
    current_row += 1

    # Milestone rows
    milestones = [
        ("Invoice 1 — 35% at Signing",        0,                   0.35),
        ("Invoice 2 — 35% at 1/3 Completion", num_months // 3,     0.35),
        ("Invoice 3 — 20% at 2/3 Completion", 2 * num_months // 3, 0.20),
        ("Invoice 4 — 10% at Completion",      num_months - 1,      0.10),
    ]
    pay_cells = []
    for i, (desc, offset, pct) in enumerate(milestones):
        r = current_row
        due = (start_dt + relativedelta(months=offset)).strftime("%B %Y")
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 15
        _style(ws.cell(row=r, column=1, value=desc), bg=bg)
        _style(ws.cell(row=r, column=2, value=due), bg=bg, align="center")
        _style(ws.cell(row=r, column=3, value=f"={pay_base}*{pct}"),
               bg=bg, align="center", num_fmt="$#,##0")
        _style(ws.cell(row=r, column=4), bg=bg)
        pay_cells.append(f"C{r}")
        current_row += 1

    # Payment total
    ws.row_dimensions[current_row].height = 15
    _style(ws.cell(row=current_row, column=1, value="Total"), bold=True, bg=TOTAL_BG)
    _style(ws.cell(row=current_row, column=2), bold=True, bg=TOTAL_BG)
    _style(ws.cell(row=current_row, column=3,
                   value="=SUM(" + ",".join(pay_cells) + ")"),
           bold=True, bg=TOTAL_BG, align="center", num_fmt="$#,##0")
    _style(ws.cell(row=current_row, column=4), bold=True, bg=TOTAL_BG)

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = f"E{WBS_HDR + 1}"

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_workplan.py config.json output.xlsx")
        sys.exit(1)
    build_workplan(sys.argv[1], sys.argv[2])

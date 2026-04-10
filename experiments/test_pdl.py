"""
PDL API diagnostic script.

Single-person mode:
    python experiments/test_pdl.py "First" "Last" "email@example.com"
    python experiments/test_pdl.py "First" "Last"   (no email)

Batch mode (CSV or Excel, up to 10 rows):
    python experiments/test_pdl.py path/to/file.csv
    python experiments/test_pdl.py path/to/file.xlsx

Prints the full raw PDL response for each record so you can audit
exactly which fields are available.
"""

import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
import requests
from enrichment import column_map

api_key = config.get_api_key()
if not api_key:
    print("ERROR: No API key found. Open the app, go to Settings, and save your PDL key.")
    sys.exit(1)

MAX_ROWS = 10


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_file(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    elif suffix in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            if all(v is None or str(v).strip() == "" for v in record.values()):
                continue
            rows.append(record)
        return rows
    else:
        print(f"ERROR: Unsupported file type '{suffix}'. Use .csv, .xlsx, or .xls.")
        sys.exit(1)


def _call_pdl(first: str, last: str, email: str = "", location: str = "") -> dict:
    payload = {"first_name": first, "last_name": last}
    if email:
        payload["email"] = email
    if location:
        payload["location"] = location
    response = requests.post(
        "https://api.peopledatalabs.com/v5/person/enrich",
        headers={"X-Api-Key": api_key},
        json=payload,
    )
    return response


def _print_result(label: str, response) -> None:
    print(f"\n{'=' * 60}")
    print(f"Record: {label}")
    print(f"HTTP Status: {response.status_code}")
    print("-" * 60)
    if response.status_code == 200:
        data = response.json()
        person = data.get("data") or data  # v5 wraps in "data"; fall back for safety
        print(f"MATCH FOUND  (likelihood: {data.get('likelihood')} / 10)\n")
        print("Fields the tool currently reads:")
        print(f"  job_company_name    : {person.get('job_company_name')}")
        print(f"  job_title           : {person.get('job_title')}")
        print(f"  job_company_industry: {person.get('job_company_industry')}")
        print(f"\nFull PDL response:\n{json.dumps(person, indent=2)}")
    elif response.status_code == 404:
        print("NO MATCH — PDL does not have a record for this person.")
    elif response.status_code == 401:
        print("UNAUTHORIZED — Check your API key in Settings.")
    elif response.status_code == 402:
        print("PAYMENT REQUIRED — PDL credits may be exhausted.")
    else:
        print(f"Unexpected response:\n{response.text}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

args = sys.argv[1:]

if not args:
    print("Usage:")
    print('  python experiments/test_pdl.py "First" "Last" "email@example.com"')
    print("  python experiments/test_pdl.py path/to/file.csv")
    print("  python experiments/test_pdl.py path/to/file.xlsx")
    sys.exit(1)

# Detect mode: file path vs inline name/email
first_arg = Path(args[0])
if first_arg.suffix.lower() in (".csv", ".xlsx", ".xls"):
    # --- Batch mode ---
    rows = _read_file(first_arg)
    if not rows:
        print("ERROR: File is empty or has no data rows.")
        sys.exit(1)

    rows = rows[:MAX_ROWS]
    headers = list(rows[0].keys())
    mapping = column_map.detect(headers)

    print(f"\nFile: {first_arg.name}  ({len(rows)} record(s) — max {MAX_ROWS})")
    print(f"Detected columns: { {k: v for k, v in mapping.items() if v} }")

    missing = [k for k, v in mapping.items() if v is None and k in ("first_name", "last_name")]
    if missing:
        print(f"\nERROR: Could not detect required columns: {missing}")
        print(f"Available headers: {headers}")
        sys.exit(1)

    for i, row in enumerate(rows, 1):
        first, last, location, email = column_map.extract_name_and_location(row, mapping)
        if not first and not last:
            print(f"\nRow {i}: skipped (no name found)")
            continue
        label = f"{first} {last}" + (f" <{email}>" if email else "")
        response = _call_pdl(first, last, email, location)
        _print_result(f"Row {i} — {label}", response)

else:
    # --- Single-person mode ---
    if len(args) < 2:
        print('Usage: python experiments/test_pdl.py "First" "Last" "email@example.com"')
        sys.exit(1)

    first, last = args[0], args[1]
    email = args[2] if len(args) >= 3 else ""

    print(f"\nQuerying PDL for: {first} {last}" + (f" <{email}>" if email else ""))
    response = _call_pdl(first, last, email)
    _print_result(f"{first} {last}", response)

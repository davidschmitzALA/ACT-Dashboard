import openpyxl


def read_attendees(filepath: str) -> list[dict]:
    """Read attendee rows from an Excel file. Returns a list of dicts with normalized keys."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    attendees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        # Skip entirely blank rows
        if all(v is None or str(v).strip() == "" for v in record.values()):
            continue
        attendees.append(record)

    return attendees

import openpyxl
from openpyxl.styles import PatternFill, Font

_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_BOLD = Font(bold=True)

ENRICHMENT_HEADERS = [
    "Employer", "Job Title", "Department", "Industry (v2)",
    "HR Flag", "C-Suite Flag", "L&D Flag", "Marketing/Comms Flag", "Match Confidence",
    "Birth Year", "LinkedIn URL", "LinkedIn Username",
    "Facebook URL", "Twitter URL",
    "Work Email", "Personal Email", "Mobile Phone",
    "Company Website", "Company Size",
    "Job Start Date", "Inferred Salary",
]


def _write_sheet(ws, original_headers: list, rows: list[dict], include_flags: bool = True):
    headers = original_headers + (ENRICHMENT_HEADERS if include_flags else [])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _BOLD

    for row in rows:
        values = [row.get(h) for h in original_headers]
        if include_flags:
            exp = row.get("experience")
            values += [
                row.get("employer"),
                row.get("job_title"),
                row.get("department"),
                row.get("job_company_industry_v2"),
                "Yes" if row.get("hr_flag") else "No",
                "Yes" if row.get("csuite_flag") else "No",
                "Yes" if row.get("ld_flag") else "No",
                "Yes" if row.get("marketing_flag") else "No",
                row.get("confidence"),
                row.get("birth_year"),
                row.get("linkedin_url"),
                row.get("linkedin_username"),
                row.get("facebook_url"),
                row.get("twitter_url"),
                row.get("work_email"),
                row.get("recommended_personal_email"),
                row.get("mobile_phone"),
                row.get("job_company_website"),
                row.get("job_company_size"),
                row.get("job_start_date"),
                row.get("inferred_salary"),
            ]
        ws.append(values)

    # Highlight flagged rows
    if include_flags:
        for i, row in enumerate(rows, start=2):
            if row.get("hr_flag") or row.get("csuite_flag") or row.get("ld_flag") or row.get("marketing_flag"):
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col_idx).fill = _YELLOW


def write_results(
    output_path: str,
    original_headers: list,
    matched: list[dict],
    unmatched: list[dict],
):
    wb = openpyxl.Workbook()

    ws_main = wb.active
    ws_main.title = "Enriched"
    _write_sheet(ws_main, original_headers, matched, include_flags=True)

    ws_unmatched = wb.create_sheet("Unmatched")
    _write_sheet(ws_unmatched, original_headers, unmatched, include_flags=False)

    wb.save(output_path)
